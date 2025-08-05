# app/scraper.py (actualizado)
import json
from openpyxl import load_workbook
from datetime import time


def parse_excel(file_path: str) -> dict:
    wb = load_workbook(file_path)
    sheet = wb.active

    data = {"days": []}
    current_day = None
    current_direction = None
    stations = []

    for row in sheet.iter_rows(values_only=True):
        if not any(row):
            continue

        if row[0] and isinstance(row[0], str) and "HORARIO" in row[0]:
            parts = row[0].split('-')
            if len(parts) > 1:
                current_day = parts[1].strip()
            if len(parts) > 2:
                direction_str = parts[2].strip()
                if "VILLA ROSA" in direction_str:
                    current_direction = "Hacia Villa Rosa"
                elif "RETIRO" in direction_str:
                    current_direction = "Hacia Retiro"
                else:
                    current_direction = None
            stations = []
            continue

        if row[0] == "N°" or (row[0] is None and any(row[1:])):
            stations = [str(cell).strip() for cell in row[1:] if cell]
            continue

        if row[0] is not None:
            if isinstance(row[0], (int, float, str)):
                try:
                    trip_id = str(int(row[0]))
                except:
                    continue

                if not current_day or not current_direction or not stations:
                    continue

                schedule = {}
                for i, station in enumerate(stations):
                    time_val = row[i + 1] if (i + 1) < len(row) else None

                    if isinstance(time_val, time):
                        schedule[station] = time_val.strftime("%H:%M")
                    elif time_val and isinstance(time_val, str):
                        schedule[station] = time_val.strip()
                    else:
                        schedule[station] = None

                day_entry = next((d for d in data["days"] if d["day_type"] == current_day), None)
                if not day_entry:
                    day_entry = {"day_type": current_day, "directions": []}
                    data["days"].append(day_entry)

                dir_entry = next((d for d in day_entry["directions"] if d["direction"] == current_direction), None)
                if not dir_entry:
                    # Guardar el orden de estaciones para esta dirección
                    dir_entry = {
                        "direction": current_direction,
                        "stations_order": stations,  # <--- NUEVO: Orden de estaciones
                        "trips": []
                    }
                    day_entry["directions"].append(dir_entry)

                dir_entry["trips"].append({
                    "trip_id": trip_id,
                    "schedule": schedule
                })

    return data


def save_json(data: dict, output_path: str):
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
